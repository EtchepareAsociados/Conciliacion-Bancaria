import os, re, json, io, datetime
from flask import Flask, render_template, request, jsonify, send_file, session
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'etchepare2026conciliacion')

USUARIOS = {
    os.environ.get('USER1_NAME', 'admin'):     os.environ.get('USER1_PASS', 'etchepare2026'),
    os.environ.get('USER2_NAME', 'compañera'): os.environ.get('USER2_PASS', 'etchepare2026b'),
}

TOLERANCIA = 10000
MESES = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO',
         'JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']

MES_BG = {'ENERO':'DBEAFE','FEBRERO':'BAE6FD','MARZO':'BBF7D0','ABRIL':'FEF08A',
           'MAYO':'FED7AA','JUNIO':'FECACA','JULIO':'E9D5FF','AGOSTO':'FBCFE8',
           'SEPTIEMBRE':'A7F3D0','OCTUBRE':'FDE68A','NOVIEMBRE':'BAE6FD','DICIEMBRE':'DDD6FE'}
MES_FG = {'ENERO':'1E3A8A','FEBRERO':'075985','MARZO':'14532D','ABRIL':'713F12',
           'MAYO':'7C2D12','JUNIO':'7F1D1D','JULIO':'4C1D95','AGOSTO':'831843',
           'SEPTIEMBRE':'064E3B','OCTUBRE':'78350F','NOVIEMBRE':'0C4A6E','DICIEMBRE':'3B0764'}

C_NAVY  = '1F3864'
C_WHITE = 'FFFFFF'
C_GRAY_H= 'F8FAFC'
C_GRAY_B= 'CBD5E1'
C_BLACK = '1E293B'
C_MUTED = '475569'

# ── Archivo de decisiones persistentes ──
DECISIONES_PATH = os.path.join(os.path.dirname(__file__), 'decisiones.json')

def load_decisiones():
    if not os.path.exists(DECISIONES_PATH):
        return {}
    with open(DECISIONES_PATH, encoding='utf-8') as f:
        return json.load(f)

def save_decisiones(d):
    with open(DECISIONES_PATH, 'w', encoding='utf-8') as f:
        json.dump(d, f, ensure_ascii=False, indent=2)

# ── Helpers ──
def extract_rut(desc):
    if not desc or pd.isna(desc): return None
    m = re.match(r'^(\d{7,10}[Kk]?)\s', str(desc))
    return m.group(1).upper() if m else None

def norm_rut(r):
    if not r: return None
    return str(r).strip().replace('.','').replace('-','').upper().lstrip('0')

def sig_mes(p):
    if not p: return ''
    u = str(p).strip().upper()
    return MESES[(MESES.index(u)+1)%12] if u in MESES else ''

# ── Load BDs ──
def load_bd():
    bd_path = os.path.join(os.path.dirname(__file__), 'bd_arrendatarios.json')
    with open(bd_path, encoding='utf-8') as f:
        records = json.load(f)
    lookup = {}
    for r in records:
        k = norm_rut(r.get('RESPONSABLE',''))
        if k:
            lookup.setdefault(k, []).append(r)
    return lookup

def load_bd_caja():
    bd_path = os.path.join(os.path.dirname(__file__), 'bd_recuperacion_caja.json')
    if not os.path.exists(bd_path):
        return {}
    with open(bd_path, encoding='utf-8') as f:
        records = json.load(f)
    lookup = {}
    for r in records:
        k = norm_rut(r.get('RESPONSABLE',''))
        if k:
            lookup.setdefault(k, []).append(r)
    return lookup

BD_LOOKUP      = load_bd()
BD_CAJA_LOOKUP = load_bd_caja()

def find_match(rut_norm, monto):
    if not rut_norm or rut_norm not in BD_LOOKUP: return None
    cands = BD_LOOKUP[rut_norm]
    for c in cands:
        if c['MONTO_ESP'] == monto:
            return {**c, 'tipo':'OK', 'diff':0}
    for c in cands:
        if c.get('MONEDA','CLP')=='UF' and abs(c['MONTO_ESP']-monto)<=TOLERANCIA:
            return {**c, 'tipo':'OK', 'diff':monto-c['MONTO_ESP']}
    best = min(cands, key=lambda c: abs(c['MONTO_ESP']-monto))
    diff = monto - best['MONTO_ESP']
    return {**best, 'tipo':'DIF+' if diff>0 else 'DIF-', 'diff':diff}

def find_match_caja(rut_norm):
    if not rut_norm or rut_norm not in BD_CAJA_LOOKUP: return None
    return BD_CAJA_LOOKUP[rut_norm][0]

# ── Parse historial ──
def parse_historial(wb):
    keys = set()
    ultimo_mes   = {}
    ultimo_monto = {}  # último monto pagado por RUT
    sheets_validas = ['ARRIENDOS','ARRIENDOS OK','ARRIENDOS CON DIFERENCIAS',
                      'REAJUSTES PENDIENTES','RESERVAS','RECUPERACIÓN CAJA']
    for sname in wb.sheetnames:
        if sname == 'RESUMEN' or sname not in sheets_validas: continue
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        data_start = 1
        for i, row in enumerate(rows[:5]):
            if row and str(row[0]).strip().upper() == 'FECHA':
                data_start = i + 1
                break
        for row in rows[data_start:]:
            if not row or not row[0]: continue
            fecha   = str(row[0]).strip() if row[0] else ''
            desc    = str(row[3]).strip() if row[3] else ''
            monto   = row[4]
            periodo = str(row[5]).strip().upper() if row[5] else ''
            rut     = norm_rut(extract_rut(desc))
            if fecha and monto:
                try: keys.add(f"{rut}|{int(float(monto))}|{fecha}")
                except: pass
            if rut and periodo in MESES:
                # Guardar siempre el mes más reciente (más avanzado en el año)
                actualizar = False
                if rut not in ultimo_mes:
                    actualizar = True
                else:
                    idx_actual = MESES.index(ultimo_mes[rut])
                    idx_nuevo  = MESES.index(periodo)
                    if idx_nuevo > idx_actual:
                        actualizar = True
                    elif idx_actual >= 10 and idx_nuevo <= 1:
                        actualizar = True
                if actualizar:
                    ultimo_mes[rut] = periodo
                    # Guardar también el monto de ese último pago
                    try:
                        ultimo_monto[rut] = int(float(monto))
                    except: pass
    return keys, ultimo_mes, ultimo_monto

# ── Parse cartola ──
def parse_cartola(wb):
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    abonos = []
    for row in rows[13:]:
        if not row or len(row) < 8: continue
        try: monto = int(float(row[0]))
        except: continue
        if str(row[7]).strip() != 'A': continue
        desc  = str(row[1]).strip() if row[1] else ''
        fecha = str(row[3]).strip() if row[3] else ''
        ndoc  = str(row[4]).strip() if row[4] else ''
        abonos.append({
            'monto': monto, 'desc': desc, 'fecha': fecha, 'ndoc': ndoc,
            'rut': extract_rut(desc), 'rut_norm': norm_rut(extract_rut(desc)),
            'idx': len(abonos)
        })
    return abonos

# ── Detectar mes cartola ──
def detectar_mes(abonos):
    for a in abonos:
        if a['fecha']:
            try: return MESES[int(a['fecha'].split('/')[1])-1]
            except: pass
    return ''

# ── Propuesta inteligente por carpeta ──
def proponer_clasificacion(carpeta_id, pagos, monto_esp, ultimo_mes_rut, mes_cartola, ultimo_monto_pagado=0):
    """
    Regla simple y clara:
    - Acumular pagos consecutivos hasta que el total >= 70% del arriendo → cerrar mes
    - EXCEPCIÓN: si hay múltiples pagos en el mismo día y cada uno >= 70% → meses distintos
    - Una vez cerrado el mes:
        · pago <= 20% → rec. caja / reajuste
        · pago >= 70% → nuevo mes
        · entre 20-70% → abono próximo mes
    - SMART OBS: si pagó igual que el mes pasado pero diferente al esperado → posible reajuste
    """
    from collections import OrderedDict

    resultado        = []
    mes_actual       = sig_mes(ultimo_mes_rut) or mes_cartola
    diff_pendiente   = 0
    mes_cerrado      = False
    acumulado_mes    = 0
    pagos_mes_actual = []

    def cerrar():
        nonlocal acumulado_mes, pagos_mes_actual, mes_actual, diff_pendiente
        total = acumulado_mes
        diff  = total - monto_esp
        if diff == 0:   estado = 'OK';                             clasificacion = 'ok'
        elif diff > 0:  estado = f'▲ DE MÁS +${diff:,.0f}';       clasificacion = 'dif'
        else:           estado = f'▼ DE MENOS -${abs(diff):,.0f}'; clasificacion = 'dif'
        diff_pendiente = abs(diff) if diff < 0 else 0
        # Smart obs: siempre inicializar primero
        obs = ''
        if len(pagos_mes_actual) > 1:
            obs = f'Pago fraccionado — total ${total:,.0f}'
        if clasificacion == 'dif' and ultimo_monto_pagado > 0:
            if abs(total - ultimo_monto_pagado) <= max(ultimo_monto_pagado * 0.02, 3000):
                obs = f'Pagó igual que mes anterior (${ultimo_monto_pagado:,.0f}) — posible reajuste pendiente'
        for p in pagos_mes_actual:
            resultado.append({**p, 'carpeta': carpeta_id,
                'mes': mes_actual, 'estado': estado,
                'clasificacion': clasificacion, 'obs': obs})
        mes_actual       = sig_mes(mes_actual) or mes_actual
        acumulado_mes    = 0
        pagos_mes_actual = []

    # Agrupar por fecha
    grupos = OrderedDict()
    for p in pagos:
        grupos.setdefault(p['fecha'], []).append(p)
    fechas = list(grupos.keys())

    for fi, fecha in enumerate(fechas):
        grupo    = grupos[fecha]
        hay_mas  = fi < len(fechas) - 1

        if not mes_cerrado:
            # Caso especial: múltiples pagos en el mismo día, cada uno >= 70% → meses distintos
            if len(grupo) > 1 and all(p['monto'] / monto_esp >= 0.70 for p in grupo):
                for pago in grupo:
                    acumulado_mes    = pago['monto']
                    pagos_mes_actual = [pago]
                    cerrar()
                mes_cerrado = True
                continue

            # Acumular este grupo al mes actual
            acumulado_mes    += sum(p['monto'] for p in grupo)
            pagos_mes_actual += grupo
            ratio = acumulado_mes / monto_esp if monto_esp > 0 else 1

            if ratio >= 1.0:
                # Pagó completo o de más → cerrar mes
                cerrar()
                mes_cerrado = True
            elif not hay_mas:
                # Último grupo → cerrar con lo que hay
                cerrar()
                mes_cerrado = True
            elif ratio >= 0.70:
                # Tenemos >= 70% pero hay más pagos
                # Ver si el próximo grupo son arriendos completos independientes
                prox_grupo = grupos[fechas[fi + 1]]
                prox_total = sum(p['monto'] for p in prox_grupo)
                prox_ratio = prox_total / monto_esp if monto_esp > 0 else 1

                if prox_ratio <= 0.20:
                    # El próximo es muy pequeño → es reajuste/caja, cerrar mes ya
                    cerrar()
                    mes_cerrado = True
                elif len(prox_grupo) > 1 and all(p['monto'] / monto_esp >= 0.70 for p in prox_grupo):
                    # El próximo son múltiples arriendos completos → cerrar mes ya
                    cerrar()
                    mes_cerrado = True
                elif prox_ratio >= 0.70 and acumulado_mes / monto_esp >= 0.90:
                    # Ya tenemos >= 90% y el próximo también es grande → cerrar mes
                    cerrar()
                    mes_cerrado = True
                # Si no, seguir acumulando (el próximo complementa este mes)
            # Si ratio < 0.70, seguir acumulando siempre

        else:
            # Mes cerrado → clasificar cada pago individualmente
            for pago in grupo:
                ratio = pago['monto'] / monto_esp if monto_esp > 0 else 1
                diff  = pago['monto'] - monto_esp

                if diff_pendiente > 0 and abs(pago['monto'] - diff_pendiente) <= max(diff_pendiente * 0.15, 5000):
                    resultado.append({**pago, 'carpeta': carpeta_id,
                        'mes': mes_actual, 'estado': '⚠️ REAJUSTE PENDIENTE',
                        'clasificacion': 'reajuste',
                        'obs': f'Cubre diferencia pendiente de ${diff_pendiente:,.0f}'})
                    diff_pendiente = 0

                elif ratio <= 0.20:
                    resultado.append({**pago, 'carpeta': carpeta_id,
                        'mes': mes_actual, 'estado': '🏠 RECUPERACIÓN CAJA',
                        'clasificacion': 'caja', 'obs': 'Monto bajo — posible recuperación caja'})

                elif ratio >= 0.70:
                    # Nuevo arriendo → acumular para siguiente mes
                    acumulado_mes    = pago['monto']
                    pagos_mes_actual = [pago]
                    mes_cerrado      = False
                    if not hay_mas:
                        cerrar()
                        mes_cerrado = True

                else:
                    resultado.append({**pago, 'carpeta': carpeta_id,
                        'mes': mes_actual,
                        'estado': f'▼ ABONO PRÓX. MES -${abs(diff):,.0f}',
                        'clasificacion': 'dif',
                        'obs': 'Abono parcial — posible pago anticipado próximo mes'})

    # Pagos que quedaron sin cerrar
    if pagos_mes_actual:
        cerrar()

    return resultado

def procesar(hist_wb, cartola_wb):
    hist_keys, ultimo_mes, ultimo_monto = parse_historial(hist_wb)
    abonos = parse_cartola(cartola_wb)
    mes_cartola = detectar_mes(abonos)

    res_ok    = []
    res_dif   = []
    res_res   = []
    res_caja  = []
    carpetas  = {}  # Para vista interactiva: {carpeta_id: {info, pagos[]}}

    for a in abonos:
        key = f"{a['rut_norm']}|{a['monto']}|{a['fecha']}"
        if key in hist_keys: continue

        base = {
            'fecha': a['fecha'], 'rut': a['rut'] or 'Sin RUT',
            'monto': a['monto'], 'desc': a['desc'][:55],
            'ndoc': a['ndoc'], 'mes': mes_cartola, 'idx': a['idx']
        }

        # ¿Es dueño / recuperación caja?
        match_caja = find_match_caja(a['rut_norm'])
        if match_caja:
            nombre = match_caja.get('NOMBRE', '')
            rec = {**base, 'carpeta': '', 'estado': 'RECUPERACIÓN CAJA', 'nombre_dueno': nombre, 'clasificacion': 'caja'}
            res_caja.append(rec)
            continue

        # ¿Es arrendatario?
        match = find_match(a['rut_norm'], a['monto'])
        if not match:
            res_res.append({**base, 'carpeta': '', 'estado': 'RESERVA', 'clasificacion': 'reserva'})
            continue

        carpeta_id = str(match['CARPETA'])
        monto_esp  = match['MONTO_ESP']

        if carpeta_id not in carpetas:
            carpetas[carpeta_id] = {
                'carpeta': carpeta_id,
                'rut': a['rut'] or 'Sin RUT',
                'rut_norm': a['rut_norm'] or '',
                'monto_esp': monto_esp,
                'ultimo_mes': ultimo_mes.get(a['rut_norm'], ''),
                'pagos': []
            }
        carpetas[carpeta_id]['pagos'].append({**base, 'carpeta': carpeta_id})

    # Ordenar pagos de cada carpeta por índice (orden cartola)
    for cid, info in carpetas.items():
        info['pagos'] = sorted(info['pagos'], key=lambda x: x['idx'])

    # Generar propuestas por carpeta
    for cid, info in carpetas.items():
        propuestas = proponer_clasificacion(
            cid, info['pagos'], info['monto_esp'],
            info['ultimo_mes'], mes_cartola,
            ultimo_monto.get(info['rut_norm'], 0)
        )
        for p in propuestas:
            if p['clasificacion'] == 'ok':
                res_ok.append(p)
            elif p['clasificacion'] == 'reajuste':
                res_dif.append({**p, 'es_reajuste': True})
            else:
                res_dif.append(p)

    res_ok   = sorted(res_ok,   key=lambda x: x['idx'])
    res_dif  = sorted(res_dif,  key=lambda x: x['idx'])
    res_res  = sorted(res_res,  key=lambda x: x['idx'])
    res_caja = sorted(res_caja, key=lambda x: x['idx'])

    # Vista interactiva — lista de carpetas con resumen
    vista_carpetas = []
    for cid, info in sorted(carpetas.items(), key=lambda x: x[0]):
        total = sum(p['monto'] for p in info['pagos'])
        diff  = total - info['monto_esp']
        propuestas = proponer_clasificacion(
            cid, info['pagos'], info['monto_esp'],
            info['ultimo_mes'], mes_cartola,
            ultimo_monto.get(info['rut_norm'], 0)
        )
        estado_prop = propuestas[0]['clasificacion'] if propuestas else 'dif'
        # Solo incluir en vista si hay algo que revisar (no las OK exactas)
        if estado_prop == 'ok' and len(info['pagos']) == 1:
            continue
        # Enriquecer cada pago con su mes y diferencia individual
        pagos_enriquecidos = []
        for p in propuestas:
            pagos_enriquecidos.append({
                **p,
                'diff_ind': p['monto'] - info['monto_esp']
            })

        vista_carpetas.append({
            'carpeta':   cid,
            'rut':       info['rut'],
            'monto_esp': info['monto_esp'],
            'total_pag': total,
            'diff':      diff,
            'n_pagos':   len(info['pagos']),
            'pagos':     pagos_enriquecidos,
            'propuesta': estado_prop,
            'mes':       sig_mes(info['ultimo_mes']) or mes_cartola,
        })

    return res_arr, res_res, res_caja, mes_cartola, vista_carpetas

# ── Escribir filas en hoja Excel ──
def escribir_filas(ws, rows, has_carpeta=True):
    last_row = ws.max_row
    while last_row > 3 and ws.cell(last_row, 1).value in [None, '']:
        last_row -= 1

    for row_data in rows:
        last_row += 1
        row_bg = C_GRAY_H if last_row % 2 == 0 else C_WHITE

        obs = row_data.get('obs', '')
        if has_carpeta:
            vals = [row_data['fecha'], row_data.get('carpeta',''),
                    row_data['rut'], row_data['desc'],
                    row_data['monto'], row_data.get('mes',''),
                    row_data['estado'], obs]
        else:
            vals = [row_data['fecha'], '',
                    row_data['rut'], row_data['desc'],
                    row_data['monto'], row_data.get('mes',''),
                    row_data['estado'], obs]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=last_row, column=col, value=val)
            c.fill      = PatternFill('solid', fgColor=row_bg)
            c.border    = Border(bottom=Side(style='thin',color='E2E8F0'),
                                 left=Side(style='thin',color='E2E8F0'),
                                 right=Side(style='thin',color='E2E8F0'))
            c.font      = Font(name='Calibri', size=9, color=C_BLACK)
            c.alignment = Alignment(vertical='center')

        ws.cell(last_row,1).alignment = Alignment(horizontal='center',vertical='center')
        ws.cell(last_row,2).font      = Font(name='Calibri',bold=True,size=10,color=C_NAVY)
        ws.cell(last_row,2).alignment = Alignment(horizontal='center',vertical='center')
        ws.cell(last_row,3).font      = Font(name='Courier New',size=9,color=C_BLACK)
        ws.cell(last_row,3).alignment = Alignment(horizontal='center',vertical='center')

        mc = ws.cell(last_row,5)
        mc.font          = Font(name='Courier New',bold=True,size=9,color=C_BLACK)
        mc.number_format = '#,##0'
        mc.alignment     = Alignment(horizontal='right',vertical='center')

        pc  = ws.cell(last_row,6)
        per = str(pc.value).strip().upper() if pc.value else ''
        if per in MES_BG:
            pc.fill = PatternFill('solid',fgColor=MES_BG[per])
            pc.font = Font(name='Calibri',bold=True,size=8,color=MES_FG.get(per,C_BLACK))
        else:
            pc.font = Font(name='Calibri',size=8,color=C_BLACK)
        pc.alignment = Alignment(horizontal='center',vertical='center')

        ec     = ws.cell(last_row,7)
        estado = str(ec.value) if ec.value else ''
        if estado == 'OK':
            ec.font = Font(name='Calibri',bold=True,size=8,color='14532D')
            ec.fill = PatternFill('solid',fgColor='DCFCE7')
        elif estado.startswith('▲'):
            ec.font = Font(name='Calibri',bold=True,size=8,color='1D4ED8')
            ec.fill = PatternFill('solid',fgColor='DBEAFE')
        elif estado.startswith('▼'):
            ec.font = Font(name='Calibri',bold=True,size=8,color='991B1B')
            ec.fill = PatternFill('solid',fgColor='FEE2E2')
        elif '⚠️' in estado:
            ec.font = Font(name='Calibri',bold=True,size=8,color='78350F')
            ec.fill = PatternFill('solid',fgColor='FEF3C7')
        elif estado == 'RECUPERACIÓN CAJA':
            ec.font = Font(name='Calibri',bold=True,size=8,color='7C2D12')
            ec.fill = PatternFill('solid',fgColor='FED7AA')
        ec.alignment = Alignment(horizontal='center',vertical='center')

        ws.cell(last_row,8).font = Font(name='Calibri',size=8,color=C_MUTED,italic=True)

def clonar_encabezado(wb, origen_name, destino_ws, nuevo_titulo):
    if origen_name not in wb.sheetnames: return
    ws_orig = wb[origen_name]
    for row_idx in range(1, 4):
        for col_idx in range(1, 9):
            src = ws_orig.cell(row=row_idx, column=col_idx)
            dst = destino_ws.cell(row=row_idx, column=col_idx)
            dst.value = src.value
            if src.has_style:
                dst.font = src.font.copy(); dst.fill = src.fill.copy()
                dst.border = src.border.copy(); dst.alignment = src.alignment.copy()
                dst.number_format = src.number_format
    destino_ws.cell(row=1, column=1).value = nuevo_titulo
    for col_idx in range(1, 9):
        col_letter = get_column_letter(col_idx)
        if col_letter in ws_orig.column_dimensions:
            destino_ws.column_dimensions[col_letter].width = ws_orig.column_dimensions[col_letter].width

def generar_excel(hist_wb, res_arr, res_res, res_caja):
    # Eliminar pestañas obsoletas
    for eliminar in ['SIN ADM', 'EFECTIVO-CHEQUE', 'ARRIENDOS OK',
                     'ARRIENDOS CON DIFERENCIAS', 'REAJUSTES PENDIENTES']:
        if eliminar in hist_wb.sheetnames:
            del hist_wb[eliminar]

    # Renombrar ARRIENDOS si existe, o crear
    if 'ARRIENDOS' not in hist_wb.sheetnames:
        ws_arr = hist_wb.create_sheet('ARRIENDOS', 1)
        clonar_encabezado(hist_wb, 'RESERVAS', ws_arr, 'Arriendos Reconocidos 2026')
    
    # Crear REAJUSTES PENDIENTES si no existe
    if 'REAJUSTES PENDIENTES' not in hist_wb.sheetnames:
        ws_rea = hist_wb.create_sheet('REAJUSTES PENDIENTES')
        clonar_encabezado(hist_wb, 'ARRIENDOS', ws_rea, 'Reajustes Pendientes 2026')

    # Ordenar pestañas
    orden = ['RESUMEN','ARRIENDOS','REAJUSTES PENDIENTES','RESERVAS','RECUPERACIÓN CAJA']
    for i, nombre in enumerate(orden):
        if nombre in hist_wb.sheetnames:
            idx_actual = hist_wb.sheetnames.index(nombre)
            hist_wb.move_sheet(nombre, offset=i - idx_actual)

    # Separar reajustes de arriendos
    res_reajuste = [r for r in res_arr if r.get('clasificacion') == 'reajuste']
    res_arr_solo = [r for r in res_arr if r.get('clasificacion') != 'reajuste']

    if res_arr_solo and 'ARRIENDOS' in hist_wb.sheetnames:
        escribir_filas(hist_wb['ARRIENDOS'], res_arr_solo, True)
    if res_reajuste and 'REAJUSTES PENDIENTES' in hist_wb.sheetnames:
        escribir_filas(hist_wb['REAJUSTES PENDIENTES'], res_reajuste, True)
    if res_res and 'RESERVAS' in hist_wb.sheetnames:
        escribir_filas(hist_wb['RESERVAS'], res_res, False)
    if res_caja and 'RECUPERACIÓN CAJA' in hist_wb.sheetnames:
        escribir_filas(hist_wb['RECUPERACIÓN CAJA'], res_caja, False)

    output = io.BytesIO()
    hist_wb.save(output)
    output.seek(0)
    return output

# ── Routes ──
@app.route('/')
def index():
    if not session.get('user'):
        return render_template('login.html')
    return render_template('index.html', user=session['user'])

@app.route('/login', methods=['POST'])
def login():
    user = request.form.get('usuario','').strip()
    pwd  = request.form.get('password','').strip()
    if USUARIOS.get(user) == pwd:
        session['user'] = user
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'msg': 'Usuario o contraseña incorrectos'})

@app.route('/logout')
def logout():
    session.clear()
    return render_template('login.html')

@app.route('/procesar', methods=['POST'])
def procesar_route():
    if not session.get('user'):
        return jsonify({'error': 'No autorizado'}), 401
    try:
        hist_file    = request.files.get('historial')
        cartola_file = request.files.get('cartola')
        if not hist_file or not cartola_file:
            return jsonify({'error': 'Faltan archivos'}), 400

        hist_wb    = load_workbook(hist_file)
        cartola_wb = load_workbook(cartola_file)

        res_arr, res_res, res_caja, mes, vista_carpetas = procesar(hist_wb, cartola_wb)

        # Cargar decisiones guardadas y aplicar a vista
        decisiones = load_decisiones()
        for c in vista_carpetas:
            key = f"{c['carpeta']}_{mes}"
            if key in decisiones:
                c['propuesta'] = decisiones[key]['clasificacion']
                c['mes']       = decisiones[key].get('mes', c['mes'])
                c['obs']       = decisiones[key].get('obs', '')

        return jsonify({
            'arr':      res_arr,
            'res':      res_res,
            'caja':     res_caja,
            'mes':      mes,
            'carpetas': vista_carpetas,
            'total':    len(res_arr) + len(res_res) + len(res_caja)
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/guardar_decision', methods=['POST'])
def guardar_decision():
    if not session.get('user'):
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data       = request.json
        carpeta    = data.get('carpeta')
        mes        = data.get('mes')
        clasif     = data.get('clasificacion')
        obs        = data.get('obs', '')
        mes_asig   = data.get('mes_asignado', mes)

        decisiones = load_decisiones()
        key = f"{carpeta}_{mes}"
        decisiones[key] = {
            'clasificacion': clasif,
            'mes':           mes_asig,
            'obs':           obs,
            'fecha':         datetime.datetime.now().isoformat()
        }
        save_decisiones(decisiones)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/descargar', methods=['POST'])
def descargar_route():
    if not session.get('user'):
        return jsonify({'error': 'No autorizado'}), 401
    try:
        hist_file = request.files.get('historial')
        data      = json.loads(request.form.get('data','{}'))
        if not hist_file:
            return jsonify({'error': 'Falta historial'}), 400

        hist_wb      = load_workbook(hist_file)
        res_arr  = data.get('arr',  [])
        res_res  = data.get('res',  [])
        res_caja = data.get('caja', [])

        output = generar_excel(hist_wb, res_arr, res_res, res_caja)

        from datetime import date
        filename = f"Historial_2026_Actualizado_{date.today().isoformat()}.xlsx"
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/health')
def health():
    return 'OK', 200

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
